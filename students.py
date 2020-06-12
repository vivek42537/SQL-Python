import pandas as pd
from pandas import ExcelWriter
from openpyxl import load_workbook
#import mysql.connector
import csv
from datetime import datetime
import numpy as np
import yaml


with open('test1.yaml', 'r') as yam:
    doc = yaml.safe_load(yam)


from sqlalchemy import create_engine
engine = create_engine('mysql+pymysql://root:password@localhost:3306/testdb')

import time
import datetime

ts = time.time()
timestamp = (datetime.datetime.fromtimestamp(ts).strftime('%Y-%m-%d %H:%M:%S'),)



df = pd.read_csv (r'students.csv')
#print (df)
# df = df.rename(columns={'Name ': 'Name'})
# df = df.rename(columns={'Country ': 'Country'})
# df = df.rename(columns={'Age ': 'Age'})
df.columns = df.columns.str.strip()
#print (df.columns)
# def AlertCat_query (df):
#     cat = df.query("Country ==  'Japan '")
#     rowIndex = df.index[cat]
#     print (rowIndex)
    #df.loc[rowIndex, 'Newbie'] = 6, 
x = doc
def AlertCat (df):
    for key,value in doc["Names"].items():
        if key == 'Vivek' :
            print('1')
            print(key)
            mask = np.column_stack([df[value[0]].str.contains(key, na=False, case=True) for col in df])
            df.loc[mask.any(axis=1), 'ALERT_CAT'] = value[1]
         

        if (key == 'Jon') :
            if (np.column_stack([df.Country.str.contains('Japan', na=False, case=False) for col in df])).any():
                print('2')
                print(key)
                x = np.logical_and([df.Country.str.contains('Wakanda ', na=False, case=False) for col in df],[df.Name.str.contains('Jon', na=False, case=False) for col in df])
                mask = np.column_stack(x)
                df.loc[mask.any(axis=1), 'ALERT_CAT'] = df.loc[mask.any(axis=1), 'ALERT_CAT'].fillna('Backup Alert')

            else:
                mask = np.column_stack([df[value[0]].str.contains(key, na=False, case=False) for col in df]) 
                df.loc[mask.any(axis=1), 'ALERT_CAT'] = df.loc[mask.any(axis=1), 'ALERT_CAT'].fillna(value[1])
            

        
        if (np.column_stack([df[value[0]].str.contains(key, na=False, case=False) for col in df])).any() :
            print('3')
            print(key)
            mask = np.column_stack([df[value[0]].str.contains(key, na=False, case=False) for col in df]) 
            # if ((key == 'metric' or 'good' or 'monitor') & (np.column_stack([df.SUMMARY.str.contains('database', na=False, case=False) for col in df]))).any():
            #     df.loc[mask.any(axis=1), 'ALERT_CAT'] = df.loc[mask.any(axis=1), 'ALERT_CAT'].fillna('Sitescope Database')
            # else:
                #df.loc[mask.any(axis=1), 'ALERT_CAT'] = value[1]
            df.loc[mask.any(axis=1), 'ALERT_CAT'] = df.loc[mask.any(axis=1), 'ALERT_CAT'].fillna(value[1])

    df.fillna('Other', inplace=True)
    print (df)

    #cat = df.isin({'Age': [33]})
    #df.loc[df.index[cat], 'Newbie'] = 'Nine'
    #print(df.Name)
    # mask = np.column_stack([df.Name.str.contains("Vivek|Bill", na=False, case=False) for col in df])
    # #print(df.loc[mask.any(axis=1)])
    # df.loc[(mask.any(axis=1)), 'Newbie'] = 6
    # if (np.column_stack([df.Name.str.contains("Jon", na=False, case=False) for col in df])).any() :
    #     mask = (np.column_stack([df.Name.str.contains('Jon', na=False, case=False) for col in df]))
    #     df.loc[mask.any(axis=1), 'Newbie'] = df.loc[mask.any(axis=1), 'Newbie'].fillna('Cool')
    # mask = (np.column_stack([df.Name.str.contains('Maria', na=False, case=False) for col in df]))
    # df.loc[mask.any(axis=1), 'Newbie'] = df.loc[mask.any(axis=1), 'Newbie'].fillna('Gay')
    # print (df)
    # print (df.Newbie)

    # a = pd.isnull(df.Newbie)
    # print (a)


AlertCat (df)


#print (df)
# mydb = mysql.connector.connect(
#     host = "localhost",
#     user = "root",
#     passwd = "password",
#     database = "testdb"
# )

# print(mydb)

# mycursor = mydb.cursor()
#mycursor.execute("ALTER TABLE people ADD id MEDIUMINT primary key NOT NULL AUTO_INCREMENT")
#now = datetime.now()
#formatted_date = now.strftime('%Y-%m-%d %H:%M:%S')
#formatted_date = now.strftime('%d/%m/%Y %H:%i:%s')
#mycursor.execute("ALTER TABLE people ADD Created DATETIME")
#mycursor.execute("""INSERT into people (Created) VALUES (STR_TO_DATE(formatted_date, '%Y-%m-%d %H:%M:%S'))""")
#mycursor.execute("INSERT INTO people (Created) VALUES (%s)", (timestamp))
#print('TIME:', formatted_date)
   
#pd.read_sql('SELECT * FROM people',mydb).to_excel('foo.xlsx')
#engine.execute('CREATE TABLE people9 (Name nvarchar(50), Country nvarchar(50), Age int)')
# df.to_sql('people9', con=engine)
# for row in df.itertuples():
#     engine.execute("INSERT INTO people9 (Name, Country, Age) VALUES (%s,%s,%s)",
#         (row.Name, 
#         row.Country,
#         row.Age))


#mydb.commit()
# mycursor.execute('CREATE TABLE people2 (Name nvarchar(50), Country nvarchar(50), Age int, updated_at TIMESTAMP NOT NULL DEFAULT NOW() ON UPDATE NOW(), created_at TIMESTAMP NOT NULL DEFAULT NOW())')

# for row in df.itertuples():
#     mycursor.execute("INSERT INTO people2 (Name, Country, Age) VALUES (%s,%s,%s)",
#         (row.Name, 
#         row.Country,
#         row.Age))

#mycursor.execute("SELECT * FROM people WHERE Name = 'Aria'")
#mycursor.execute("ALTER TABLE people DROP test_date")
# book = load_workbook('foo.xlsx')
# writer = pd.ExcelWriter('foo.xlsx', engine='openpyxl')
# writer.book = book
# writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

# pd.read_sql('SELECT * FROM people',mydb).to_excel(writer, sheet_name = 'RawData')
# writer.save()
#mydb.commit()
# operation = ""
# while(operation != 'done'):
#     operation = input("Press 1, 2, 3, or 4 for correspongind CRUD operation, OR press 5 to export to excel file: ")
#     if (operation == '1'): #CREATE -insert row
#         nam = input("Name: ")
#         place = input("Country: ")
#         num = input("Age: ")
#         #sqlFormula = "INSERT INTO people_info (Name, Country, Age) VALUES(%s,%s,%s)"
#         values = [(nam, place, num)]
#         mycursor.execute("INSERT INTO people (Name, Country, Age, Created) VALUES(%s,%s,%s,%s)",(nam, place, int(num),timestamp[0]))
        
#         mydb.commit()

#     elif (operation == '2'): #READ -fetch data
#         nam = input("Name: ")
#         sql = "SELECT * FROM people WHERE Name = '{0}'"
#         query = sql.format(str(nam))
#         mycursor.execute(query)
#         myresult = mycursor.fetchall() #gets specific data you want can use 'fetchone' if you want just a single entry
#         print("sugoi")
#         for row in myresult:
#             print(row)

#     elif (operation == '3'): #UPDATE - replace data
#         nam = input("Name: ")
#         age = input("Age: ")
#         #place = input("Country:")
#         mycursor.execute("UPDATE people SET Age = %s WHERE Name = %s",(age, nam))
#         mycursor.execute("UPDATE people SET Created = %s WHERE Name = %s",(timestamp[0], nam))
#         mydb.commit()

#     elif (operation == '4'): #DELETE - delete data
#         nam = input("Name: ")
#         sql = "DELETE FROM people WHERE Name = '{0}'"
#         query = sql.format(str(nam))
#         mycursor.execute(query)
#         mydb.commit()
    
#     elif (operation == '5'): #EXPORT to excel file
#         fileName = input("File name: ")
#         sheet = input("Sheet name: ")
#         book = load_workbook(fileName)
#         writer = pd.ExcelWriter(fileName, engine='openpyxl')
#         writer.book = book
#         writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
#         pd.read_sql('SELECT * FROM people',mydb).to_excel(writer, sheet_name = sheet)
#         writer.save()

# sqlFormula = "INSERT INTO people_info (Name, Country, Age) VALUES(%s,%s,%s)"
# values = ("zeebo", "poland", 98)
# mycursor.execute(sqlFormula, values)
# mydb.commit

