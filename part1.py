import pandas as pd
from pandas import DataFrame
from pandas import ExcelWriter
from openpyxl import load_workbook
import mysql.connector
import csv
import logging

import yaml

with open('test2.yaml', 'r') as yam:
    doc = yaml.safe_load(yam)

hostInfo = doc["DatabaseInfo"]["host"]
userInfo = doc["DatabaseInfo"]["user"]
passwdInfo = doc["DatabaseInfo"]["passwd"]
databaseInfo = doc["DatabaseInfo"]["database"]

# import sqlalchemy
# from sqlalchemy import create_engine
# engine = create_engine('mysql+pymysql://root:password@localhost:3306/test2')
#fileName = input("File Name: ")
#df = pd.read_excel (r'Book3.xlsx')

logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)
formatter = logging.Formatter('%(levelname)s:%(name)s\n:%(message)s')
fileHandler = logging.FileHandler('two.log')
fileHandler.setFormatter(formatter)
logger.addHandler(fileHandler)

# logging.basicConfig(filename = 'two.log', level = logging.DEBUG, format = '%(levelname)s:%(name)s\n:%(message)s')



fileName = input("File name: ")
df = pd.read_excel (fileName)
df = df.where((pd.notnull(df)), None)
df['Created'] = df['Created'].dt.strftime('%Y-%m-%d %H:%M:%S')
df['Updated'] = df['Updated'].dt.strftime('%Y-%m-%d %H:%M:%S')
df['Resolved_at'] = df['Resolved_at'].dt.strftime('%Y-%m-%d %H:%M:%S')
df['Reassignment_count'] = df['Reassignment_count'].apply(str)

# df['Resolve_time'] = df['Resolve_time'].apply(str)
#print (df.dtypes)
# df['Book3'] = pd.to_datetime(df['Created'],unit='ms')
# df.to_excel('new_Book3.xlsx', index=False)

logger.info(df)

df.columns = df.columns.str.strip()
logger.debug(df.columns)

mydb = mysql.connector.connect(
    host = hostInfo,
    user = userInfo,
    passwd = passwdInfo,
    database = databaseInfo
)
# print (mydb)
mycursor = mydb.cursor()


# # #df.to_sql('rawData', mydb, if_exists='replace', index = False)

# mycursor.execute("SELECT * FROM RawPeople")
# myresult = mycursor.fetchall()
# #myresult2 = mycursor.keys()
# df = DataFrame(myresult, columns=['Number', 'Summary', 'Configuration_item', 'Created', 'Company', 'Assignment_group', 'Reassignment_count', 'Priority', 'Status', 'State', 'Assigned_to', 'Caller', 'Updated', 'Category', 'Category_u_category', 'Resolve_time', 'Resolved_By', 'Resolved_at', 'Resolved_by2', 'Created_by', 'Created_date', 'Location'])
#df.columns = myresult2
#print (df)

# def POD (row):
#     if row['Assignment_group'] == 'Dedicated Ops US Only – Compute' or row['Assignment_group'] == 'Sweep - Dedicated Ops_Non-Managed' or row['Assignment_group'] == 'Dedicated Ops US Only – Security' or row['Assignment_group'] == 'Dedicated Ops US Only – Network' or row['Assignment_group'] == 'Sweep - Dedicated Ops_All' :
#         return 'Dedicated POD'
    
#     elif row['Assignment_group'] == 'Sweep - FTS POD_Non-Managed' or row['Assignment_group'] == 'FTS POD – Network' or row['Assignment_group'] == 'FTS POD – Compute' or row['Assignment_group'] == 'Sweep - FTS POD_All' or row['Assignment_group'] == 'FTS POD – Security' :
#         return 'FTS POD'

#     elif row['Assignment_group'] == 'Sweep - Shared POD_Non-Managed' or row['Assignment_group'] == 'Shared POD – Network' or row['Assignment_group'] == 'Shared POD – Compute' or row['Assignment_group'] == 'Shared POD – Storage' or row['Assignment_group'] == 'Shared POD – Transport' :
#         return 'Shared POD'

def POD (row):
    for key,value in doc["POD"].items():
        if row['Assignment_group'] == key :
            return value
    


#print(df.apply (lambda row: POD(row), axis=1))
# df['POD'] = df.apply (lambda row: POD(row), axis=1)
# logger.info('Now with POD column')
# logger.info(df)

def AlertCat (row):
    for key,value in doc["Configuration Item"].items():
        if (row['Configuration_item'] is not None) and (key.lower() in row['Configuration_item'].lower()) :
            return value
    
    for key,value in doc["Summary"].items():
        if (row['Summary'] is not None) and (key.lower() in row['Summary'].lower()) :
            return value
    
    for key,value in doc["Category_u_category"].items():
        if (row['Category_u_category'] is not None) and (key.lower() in row['Category_u_category'].lower()) :
            return value

    for key,value in doc["Assignment_group"].items():
        if (row['Assignment_group'] is not None) and (key.lower() in row['Assignment_group'].lower()) :
            return value

    for key,value in doc["Company"].items():
        if (row['Company'] is not None) and (key.lower() in row['Company'].lower()) :
            return value

# def AlertCat (row):
#     if (row['Configuration_item'] is not None) and ('crlps' in row['Configuration_item'] or 'phlps' in row['Configuration_item'] or 'rchps' in row['Configuration_item'] or 'cpedellps' in row['Configuration_item']) :
#             return 'Ciena Carrier Ethernet'

#     elif 'Node Down' in row['Summary'] or 'Node or Connection Down' in row['Summary'] :
#         return 'Node Down'

#     elif 'Link Down' in row['Summary'] or 'Link Flap' in row['Summary'] or 'LINK FLAP' in row['Summary'] :
#         return 'Link Down/Link Flapping'

#     elif 'server reboot' in row['Summary'] :
#         return 'Server Reboot'
    
#     elif 'System Reboot' in row['Summary'] :
#         return 'System Reboot'

#     elif 'Agent Health Problem' in row['Summary'] :
#         return 'Agent Health Problem'

#     elif 'Agent Restart' in row['Summary'] :
#         return 'Agent Restart'
    
#     elif 'CPU uti' in row['Summary'] :
#         return 'CPU Utilization'
    
#     elif 'Memory uti' in row['Summary'] :
#         return 'Memory Utilization'

#     elif 'SWAP uti' in row['Summary'] :
#         return 'SWAP Utilization'

#     elif 'filesystem' in row['Summary'] or 'FileSystem' in row['Summary'] :
#         return 'Filesystem Utilization'
    
#     elif 'Menzies island group connection issues' in row['Summary'] :
#         return 'Menzies island group connection issues'
    
#     elif (row['Category_u_category'] is not None) and ('Security' in row['Category_u_category']) :
#         return 'Security Alert'

#     elif (row['Assignment_group'] is not None) and ('Backup' in row['Assignment_group']) :
#         return 'Backup Alert'
    
#     elif (row['Category_u_category'] is not None) and ('Storage' in row['Category_u_category']) and ('backup' in row['Summary']) :
#         return 'Backup Alert'

#     elif (row['Assignment_group'] is not None) and ('televault' in row['Assignment_group'] or 'Televault' in row['Assignment_group']) :
#         return 'Televault Alert'

#     elif (row['Category_u_category'] is not None) and ('storage' in row['Category_u_category'] or 'Storage' in row['Category_u_category'])  :
#         return 'Storage Alert'
    
#     elif 'Trap rate' in row['Summary'] :
#         return 'Trap Storm'

#     elif 'BGP' in row['Summary'] :
#         return 'BGP'

#     elif 'syslog' in row['Summary'] :
#         return 'Syslog Message'

#     elif 'ACIH' in row['Summary'] or 'ACIHost' in row['Summary']:
#         return 'ACI'
    
#     elif (row['Configuration_item'] is not None) and 'DCO' in row['Configuration_item']:
#         return 'DCO'
    
#     elif (row['Configuration_item'] is not None) and 'ipdu' in row['Configuration_item']:
#         return 'IPDU'
    
#     elif ('smti' in row['Summary'])  or ('datacenter' in row['Summary']) or ('vsphere' in row['Summary']) or ('SMTI' in row['Summary']):
#         return 'Vcenter/CO'
    
#     elif (row['Configuration_item'] is not None) and (('esx' in  row['Configuration_item']) or ('prx' in  row['Configuration_item'])) :
#             return 'Vcenter/CO'
    
#     elif 'ospf' in row['Summary'] :
#         return 'OSPF Alert'
    
#     elif 'KEEP ALIVE' in row['Summary'] :
#         return 'KEEP ALIVE'

#     elif ('Metric' in row['Summary']):
#         if ('Database' in row['Summary']):
#             return 'Sitescope Database'
#         else :
#             return 'Sitescope Others'
    
#     elif ('Good' in row['Summary']):
#         if ('Database' in row['Summary']):
#             return 'Sitescope Database'
#         else :
#             return 'Sitescope Others'
    
#     elif ('Monitor' in row['Summary']):
#         if ('Database' in row['Summary']):
#             return 'Sitescope Database'
#         else :
#             return 'Sitescope Others'

#     elif (row['Configuration_item'] is not None) and 'r2c' in row['Configuration_item']:
#         return 'R2C SR'
    
#     elif (row['Configuration_item'] is not None) and '*r2c*srm' in row['Configuration_item']:
#         return 'R2C SRM'
    
#     elif (row['Assignment_group'] is not None) and 'Transport' in row['Assignment_group'] :
#         return 'Ciena Alert'
    
#     elif ('ucs' in row['Summary']) or ('cloud platform' in row['Company']) :
#         return 'UCS ALERT'
    
#     elif (row['Configuration_item'] is not None) and ('ucs' in row['Configuration_item']) :
#         return 'transport'

#     elif ('Remote site' in row['Summary']) :
#         return 'Remote Site Unreachable'
    
#     elif (row['Configuration_item'] is not None) and ('lb-' in row['Configuration_item']) :
#         return 'LB Alert'
    
#     elif ('service' in row['Summary']) :
#         return 'Service Down'
    
#     elif ('process' in row['Summary']) :
#         return 'Process Down'

#     elif (row['Configuration_item'] is not None) and ('nsx' in row['Configuration_item']) :
#         return 'NSX Alert'
    
#     elif ('MAX0902' in row['Summary']) :
#         return 'AS/400'
    
#     elif ('Chassis down' in row['Summary']) :
#         return 'Chassis down'
    
#     elif ('Event Storm' in row['Summary']) :
#         return 'Event Storm'
    
#     elif ('FRU' in row['Summary']) :
#         return 'FRU Alert'
    
#     elif ('Forwarder' in row['Summary']) :
#         return 'Forwarder Alert'
    
#     elif ('Symmetrix' in row['Summary']) :
#         return 'Symmetrix Alert'
    
#     elif ('TVault' in row['Summary']) :
#         return 'TVault Backup Job Alert'
    
#     elif ('QA Probe' in row['Summary']) or ('ICMP Jitter' in row['Summary']) :
#         return 'ICMP Jitter'
    
#     elif (row['POD'] is not None) and ('storage' in row['POD']) and (row['POD'] == 'Shared Pod') :
#         return 'Storage Alert'
    
#     elif (row['Assignment_group'] is not None) and 'GCSM-RS-CR-OPS-SUPPORT' in row['Assignment_group'] :
#         return 'R2C SR'
    
#     elif ('Multiple events received for tenant' in row['Summary']) :
#         return 'ACI'


#print(df.apply (lambda row: AlertCat(row), axis=1))
# df['ALERT_CATEGORY'] = df.apply (lambda row: AlertCat(row), axis=1)
# logger.info('Now with ALERT_CATEGORY column')
# logger.info(df)

operation = ""
while (operation != 'done'):
    operation = input("Press 1, 2, 3, or 4 for to Create table in SQL, Insert data into SQL, ")

    if (operation == '1'): #CREATE TABLE
        mycursor.execute('CREATE TABLE RawPeople (Number nvarchar(50), Summary LONGTEXT, `Configuration_item` LONGTEXT, Created nvarchar(50), Company LONGTEXT, `Assignment_group` nvarchar(50), `Reassignment_count` nvarchar(50), Priority nvarchar(50), Status nvarchar(50), State nvarchar(50), `Assigned_to` nvarchar(50), Caller nvarchar(50), Updated nvarchar(50), Category nvarchar(50), `Category_u_category` nvarchar(50), `Resolve_time` nvarchar(50), `Resolved_By` nvarchar(50), `Resolved_at` nvarchar(50), `Resolved_by2` nvarchar(50), `Created_by` nvarchar(50), `Created_date` nvarchar(50), Location nvarchar(50))')
        mydb.commit()
    
    elif (operation == '2'): #INSERT DATA INTO SQL
        for row in df.itertuples():
            mycursor.execute("INSERT INTO RawPeople (Number, Summary, Configuration_item, Created, Company, Assignment_group, Reassignment_count, Priority, Status, State, Assigned_to, Caller, Updated, Category, Category_u_category, Resolve_time, Resolved_By, Resolved_at, Resolved_by2, Created_by, Created_date, Location) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (row.Number, 
            row.Summary, 
            row.Configuration_item,
            row.Created, 
            row.Company, 
            row.Assignment_group, 
            row.Reassignment_count, 
            row.Priority, 
            row.Status, 
            row.State, 
            row.Assigned_to, 
            row.Caller, 
            row.Updated, 
            row.Category, 
            row.Category_u_category,
            row.Resolve_time, 
            row.Resolved_By, 
            row.Resolved_at, 
            row.Resolved_by2, 
            row.Created_by, 
            row.Created_date, 
            row.Location))
        mydb.commit()
    
    elif (operation == '3'): #TAKE SQL DATA AND PUT INTO PANDAS
        mycursor.execute("SELECT * FROM RawPeople")
        myresult = mycursor.fetchall()
        df = DataFrame(myresult, columns=['Number', 'Summary', 'Configuration_item', 'Created', 'Company', 'Assignment_group', 'Reassignment_count', 'Priority', 'Status', 'State', 'Assigned_to', 'Caller', 'Updated', 'Category', 'Category_u_category', 'Resolve_time', 'Resolved_By', 'Resolved_at', 'Resolved_by2', 'Created_by', 'Created_date', 'Location'])
        logger.info('This is the data from SQL')
        logger.info(df)
    
    elif (operation == '4'): #MAP POD
        df['POD'] = df.apply (lambda row: POD(row), axis=1)
        logger.info('Now with POD column')
        logger.info(df)
        #print (df)

    elif (operation == '5'): #MAP ALERT CATEGORY
        df['ALERT_CATEGORY'] = df.apply (lambda row: AlertCat(row), axis=1)
        logger.info('Now with ALERT_CATEGORY column')
        logger.info(df)
        print (df)
    
    elif (operation == '6'): #EXPORT TO EXCEL SHEET
        fileName = input("File name: ")
        sheet = input("Sheet name: ")

        writer = pd.ExcelWriter(fileName, engine='xlsxwriter')
        df.to_excel(writer, sheet_name= sheet)
        writer.save()   

# df.to_sql('people13', con=engine)
# for row in df.itertuples():
#     mycursor.execute("INSERT INTO RawPeople (Number, Summary, Configuration_item, Created, Company, Assignment_group, Reassignment_count, Priority, Status, State, Assigned_to, Caller, Updated, Category, Category_u_category, Resolve_time, Resolved_By, Resolved_at, Resolved_by2, Created_by, Created_date, Location) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
#         (row.Number, 
#         row.Summary, 
#         row.Configuration_item,
#         row.Created, 
#         row.Company, 
#         row.Assignment_group, 
#         row.Reassignment_count, 
#         row.Priority, 
#         row.Status, 
#         row.State, 
#         row.Assigned_to, 
#         row.Caller, 
#         row.Updated, 
#         row.Category, 
#         row.Category_u_category,
#         row.Resolve_time, 
#         row.Resolved_By, 
#         row.Resolved_at, 
#         row.Resolved_by2, 
#         row.Created_by, 
#         row.Created_date, 
#         row.Location))
#df.to_sql('people13', con=engine)
mydb.commit()

# query = "INSERT INTO dataRaw (Number, Summary, Configuration item(cmdb_ci), Created, Company, Assignment group, Reassignment count, Priority, Status, State, Assigned to, Caller, Updated, Category, Category(u_category), Resolve time, Resolved By, Resolved at, Resolved by2, Created by, Created date, Location) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"

# for r in range(1, sheet.nrows):
#     Number = sheet.cell(r,0).value 
#     Summary = sheet.cell(r,1).value 
#     'Configuration item(cmdb_ci)'  = sheet.cell(r,2).value 
#     Created = sheet.cell(r,3).value 
#     Company = sheet.cell(r,4).value 
#     'Assignment group' = sheet.cell(r,5).value 
#     'Reassignment count' = sheet.cell(r,6).value 
#     Priority = sheet.cell(r,7).value 
#     Status = sheet.cell(r,8).value 
#     State = sheet.cell(r,9).value 
#     'Assigned to' = sheet.cell(r,10).value 
#     Caller = sheet.cell(r,11).value 
#     Updated = sheet.cell(r,12).value 
#     Category = sheet.cell(r,13).value 
#     Category(u_category) = sheet.cell(r,14).value 
#     'Resolve time' = sheet.cell(r,15).value 
#     'Resolved By' = sheet.cell(r,16).value 
#     'Resolved at' = sheet.cell(r,17).value 
#     'Resolved by2' = sheet.cell(r,18).value 
#     'Created by' = sheet.cell(r,19).value 
#     'Created date' = sheet.cell(r,20).value 
#     Location = sheet.cell(r,21).value 

#     values = (Number, Summary, Configuration item(cmdb_ci), Created, Company, Assignment group, Reassignment count, Priority, Status, State, Assigned to, Caller, Updated, Category, Category(u_category), Resolve time, Resolved By, Resolved at, Resolved by2, Created by, Created date, Location)
#     mycursor.execute(query,values)
#     mydb.commit()

# fileName = input("File name: ")
# sheet = input("Sheet name: ")

# writer = pd.ExcelWriter('RawData2.xlsx', engine='xlsxwriter')
# df.to_excel(writer, sheet_name='SNowProd_Omi_RD')
# writer.save()     
